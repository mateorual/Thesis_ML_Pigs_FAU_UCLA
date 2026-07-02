"""
Consolidate extracted features into a single Excel file with color-coded headers

This script:
1. Reads all extracted feature pickle files
2. Combines them with metadata (Subject, Treatment, Week, Cycles)
3. Creates a single Excel file with one row per audio
4. Color-codes column headers by source extractor
"""

import pandas as pd
from pathlib import Path
import pickle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
from tqdm import tqdm

# Feature extractor color mapping (RGB hex colors)
EXTRACTOR_COLORS = {
    'a_gat': 'FFE6E6',      # Light red
    'b_dfg': 'E6F3FF',      # Light blue
    'c_temporal': 'E6FFE6', # Light green
    'd_spectral': 'FFF9E6', # Light yellow
    'e1_mfcc': 'FFE6F9',    # Light pink
    'e2_gfcc': 'F0E6FF',    # Light purple
    'e3_plpcc': 'E6FFFF',   # Light cyan
    'e4_rastacc': 'FFE6D9', # Light orange
    'e5_pncc': 'F9FFE6',    # Light lime
    'f_lpc': 'FFE6EC',      # Light rose
    'g1_pert': 'E6E6FF',    # Light lavender
    'g2_ms': 'FFE6CC',      # Light peach
    'g3_comp_fix': 'E6F9FF',# Light sky
    'h_imf': 'F3E6FF',      # Light violet
    'i_ratio': 'E6FFE6',              # Light mint
    'z_ucla_parameters': 'FFF3CD',   # Light gold
    'z2_schlegel21_parameters': 'D4EDDA'  # Light green-gold
}

# Parameter descriptions for documentation
PARAMETER_DESCRIPTIONS = {
    # A_GAT parameters
    'f0_mean': 'Mean fundamental frequency (Hz)',
    'f0_std': 'Standard deviation of fundamental frequency',
    'f0_min': 'Minimum fundamental frequency',
    'f0_max': 'Maximum fundamental frequency',
    'jitter_local': 'Local jitter - cycle-to-cycle F0 variation (%)',
    'jitter_rap': 'Relative average perturbation of F0',
    'jitter_ppq5': 'Five-point period perturbation quotient',
    'jitter_ddp': 'Average absolute difference of differences of periods',
    'shimmer_local': 'Local shimmer - cycle-to-cycle amplitude variation (%)',
    'shimmer_apq3': 'Three-point amplitude perturbation quotient',
    'shimmer_apq5': 'Five-point amplitude perturbation quotient',
    'shimmer_apq11': 'Eleven-point amplitude perturbation quotient',
    'shimmer_dda': 'Average absolute difference of differences of amplitudes',
    'hnr': 'Harmonics-to-noise ratio (dB)',
    'nhr': 'Noise-to-harmonics ratio',
    'snr': 'Signal-to-noise ratio (dB)',

    # B_DFG parameters
    'cpp': 'Cepstral peak prominence (dB)',
    'cpps': 'Smoothed cepstral peak prominence (dB)',
    'h1_h2': 'Difference between first and second harmonic amplitudes',
    'h1_a1': 'Difference between H1 and first formant amplitude',
    'h1_a2': 'Difference between H1 and second formant amplitude',
    'h1_a3': 'Difference between H1 and third formant amplitude',

    # C_TEMPORAL parameters
    'amplitude_mean': 'Mean amplitude of signal',
    'amplitude_std': 'Standard deviation of amplitude',
    'energy_mean': 'Mean energy of signal',
    'energy_std': 'Standard deviation of energy',
    'zcr_mean': 'Mean zero-crossing rate',
    'zcr_std': 'Standard deviation of zero-crossing rate',

    # D_SPECTRAL parameters
    'spectral_centroid_mean': 'Mean spectral centroid - center of mass of spectrum (Hz)',
    'spectral_centroid_std': 'Standard deviation of spectral centroid',
    'spectral_bandwidth_mean': 'Mean spectral bandwidth - spread around centroid (Hz)',
    'spectral_bandwidth_std': 'Standard deviation of spectral bandwidth',
    'spectral_contrast_mean': 'Mean spectral contrast - peak-valley difference in subbands',
    'spectral_contrast_std': 'Standard deviation of spectral contrast',
    'spectral_flatness_mean': 'Mean spectral flatness - tonality measure (0=tonal, 1=noisy)',
    'spectral_flatness_std': 'Standard deviation of spectral flatness',
    'spectral_rolloff_mean': 'Mean spectral rolloff - frequency below which 85% energy (Hz)',
    'spectral_rolloff_std': 'Standard deviation of spectral rolloff',
    'spectral_flux_mean': 'Mean spectral flux - frame-to-frame spectral change',
    'spectral_flux_std': 'Standard deviation of spectral flux',

    # E_CEPSTRAL parameters (MFCC, GFCC, PLP, RASTA, PNCC)
    'mfcc': 'Mel-frequency cepstral coefficient',
    'gfcc': 'Gammatone-frequency cepstral coefficient',
    'plpcc': 'Perceptual linear prediction cepstral coefficient',
    'rastacc': 'RASTA-PLP cepstral coefficient (filtered)',
    'pncc': 'Power-normalized cepstral coefficient',
    '_delta': 'First derivative (velocity) of coefficient',
    '_delta2': 'Second derivative (acceleration) of coefficient',
    '_mean': 'Mean value across frames',
    '_std': 'Standard deviation across frames',

    # F_LPC parameters
    'lpc': 'Linear prediction coefficient',
    'lsf': 'Line spectral frequency',
    'lpcc': 'Linear prediction cepstral coefficient',
    'formant': 'Formant frequency (vocal tract resonance)',
    'bandwidth': 'Formant bandwidth',

    # G_AVCA parameters
    'pert': 'Perturbation measure from AVCA toolkit',
    'ms': 'Modulation spectrum feature from AVCA toolkit',
    'comp': 'Complexity measure from AVCA toolkit',
    'rpde': 'Recurrence period density entropy',
    'dfa': 'Detrended fluctuation analysis',
    'ppe': 'Pitch period entropy',

    # H_IMF parameters
    'imf': 'Intrinsic mode function feature from EMD decomposition',
    'emd': 'Empirical mode decomposition feature',

    # I_RATIO parameters
    'alpha_ratio': 'Ratio of energy 50-1000Hz to 1-5kHz (voice quality)',
    'hammar_index': 'Hammarberg index - spectral slope measure',
    'spectral_slope': 'Overall spectral slope (energy decay with frequency)',
    'tilt': 'Spectral tilt - balance of low vs high frequency energy'
}

# Extractor full descriptions
EXTRACTOR_FULL_DESCRIPTIONS = {
    'a_gat': {
        'name': 'GAT (Glottal Analysis Toolkit)',
        'description': 'Extracts fundamental frequency (F0), perturbation measures (jitter, shimmer), and noise measures (HNR, SNR). Based on glottal source analysis.',
        'reference': 'Praat-based voice quality measures'
    },
    'b_dfg': {
        'name': 'DFG (Voice Quality Features)',
        'description': 'Extracts cepstral peak prominence (CPP/CPPS) and harmonic amplitude differences (H1-H2, H1-A1, etc.). Measures voice quality and breathiness.',
        'reference': 'Hillenbrand et al., NCVS protocols'
    },
    'c_temporal': {
        'name': 'Temporal Features',
        'description': 'Extracts time-domain features including amplitude statistics, energy measures, and zero-crossing rate. Captures signal dynamics.',
        'reference': 'Standard time-domain analysis'
    },
    'd_spectral': {
        'name': 'Spectral Features',
        'description': 'Extracts frequency-domain features including spectral centroid, bandwidth, contrast, flatness, rolloff, and flux. Characterizes spectral shape.',
        'reference': 'Librosa spectral features'
    },
    'e1_mfcc': {
        'name': 'MFCC (Mel-Frequency Cepstral Coefficients)',
        'description': 'Extracts MFCCs which model human auditory perception using mel-scale filterbank. Includes deltas (velocity) and delta-deltas (acceleration).',
        'reference': 'Davis & Mermelstein (1980)'
    },
    'e2_gfcc': {
        'name': 'GFCC (Gammatone-Frequency Cepstral Coefficients)',
        'description': 'Uses gammatone filterbank that better models cochlear frequency response. More robust to noise than MFCC.',
        'reference': 'Shao et al. (2007)'
    },
    'e3_plpcc': {
        'name': 'PLPCC (Perceptual Linear Prediction)',
        'description': 'Combines spectral analysis with LP modeling, using psychoacoustic principles (equal-loudness, intensity-loudness power law).',
        'reference': 'Hermansky (1990)'
    },
    'e4_rastacc': {
        'name': 'RASTA-PLP',
        'description': 'RASTA filtering applied to PLP features. Removes slow channel variations and fast modulations, improving robustness to channel effects.',
        'reference': 'Hermansky & Morgan (1994)'
    },
    'e5_pncc': {
        'name': 'PNCC (Power-Normalized Cepstral Coefficients)',
        'description': 'Uses power-law nonlinearity and noise suppression. More robust to additive noise and channel distortion than MFCC.',
        'reference': 'Kim & Stern (2016)'
    },
    'f_lpc': {
        'name': 'LPC (Linear Prediction Coding)',
        'description': 'Models vocal tract as all-pole filter. Extracts LPC coefficients, LSF (line spectral frequencies), LPCC, and formant frequencies.',
        'reference': 'Markel & Gray (1976)'
    },
    'g1_pert': {
        'name': 'AVCA Perturbation Set',
        'description': 'Advanced perturbation measures from AVCA-ByO toolkit. Includes jitter/shimmer variants and their derivatives.',
        'reference': 'Arias-Londoño et al., AVCA-ByO toolkit'
    },
    'g2_ms': {
        'name': 'AVCA Modulation Spectrum',
        'description': 'Analyzes temporal modulations in the spectral envelope. Captures speech rhythm and articulation patterns.',
        'reference': 'Arias-Londoño et al., AVCA-ByO toolkit'
    },
    'g3_comp_fix': {
        'name': 'AVCA Complexity Features',
        'description': 'Nonlinear dynamics measures including correlation dimension, entropy measures, and RPDE. Captures signal complexity.',
        'reference': 'Arias-Londoño et al., AVCA-ByO toolkit'
    },
    'h_imf': {
        'name': 'IMF (Intrinsic Mode Functions)',
        'description': 'Features from Empirical Mode Decomposition (EMD). Decomposes signal into IMFs representing different oscillatory modes.',
        'reference': 'Huang et al. (1998), VoiceAnalysisToolbox'
    },
    'i_ratio': {
        'name': 'Spectral Ratios',
        'description': 'Ratio-based spectral features including alpha ratio, Hammarberg index, and spectral slope. Measures spectral balance.',
        'reference': 'Hammarberg et al. (1980)'
    },
    'z_ucla_parameters': {
        'name': 'UCLA Parameters (F0, Jitter, Shimmer, Q50, Flux)',
        'description': 'Five acoustic parameters computed using signal-level methods: fundamental frequency via autocorrelation, jitter and shimmer via peak detection, Q50 as the 50th-percentile spectral centroid, and spectral flux via STFT frame differences.',
        'reference': 'UCLA Voice Analysis Protocol (unified_audio_metrics_analysis_reference.py)'
    },
    'z2_schlegel21_parameters': {
        'name': 'Schlegel21 Parameters (Q50_2, Q50_10, Q50_min, PF, Q25, Dur, HNR)',
        'description': 'Seven acoustic parameters from pig vocalization literature: Q50 at call start/end and minimum across partial windows (Tallet 2013), peak frequency from whole-signal spectrum (Tallet 2013), per-frame Q25 averaged (Garcia 2015), signal duration, and HNR via Praat cross-correlation (Linhart 2015). Minimum pitch adjusted to 50 Hz for pigs.',
        'reference': 'Schlegel et al. (2021) supplementary scripts FileS1.m and FileS2'
    }
}


def load_features_from_pickle(pickle_file):
    """Load features from a pickle file"""
    try:
        with open(pickle_file, 'rb') as f:
            df = pickle.load(f)
        return df
    except Exception as e:
        print(f"  Warning: Could not load {pickle_file.name}: {e}")
        return None


def get_num_cycles(cycles_folder, squeal_name, treatment_type, subject, week):
    """
    Get number of cycles for an audio file using mirrored directory structure

    Parameters:
    -----------
    cycles_folder : Path
        Base cycles folder (Detected_Cycles/)
    squeal_name : str
        Audio file name (e.g., "ZOOM0007_LR-0001_sound_1020")
    treatment_type : str
        Treatment code: "B", "S", or "U"
    subject : str
        Subject name (e.g., "Beck and Kurt")
    week : int
        Week number

    Returns:
    --------
    int or None : Number of cycles if found, None otherwise
    """
    from svfel.utils import io

    # Treatment mapping
    TREATMENT_NAMES = {
        'B': 'Bilateral',
        'S': 'Scar',
        'U': 'Unilateral'
    }

    # Get treatment folder name
    treatment_folder = TREATMENT_NAMES.get(treatment_type)
    if not treatment_folder:
        return None

    treatment_path = cycles_folder / treatment_folder
    if not treatment_path.exists():
        return None

    # Expected folder name pattern components
    expected_week = f"week_{week}_"
    expected_subject = subject
    expected_type = f"{treatment_type}_Snip"

    # Scan all week folders in treatment directory to find matching folder
    for week_folder in treatment_path.iterdir():
        if not week_folder.is_dir():
            continue

        folder_name = week_folder.name.lower()

        # Check if folder name contains all expected components
        if (expected_week.lower() in folder_name and
            expected_subject.lower() in folder_name and
            expected_type.lower() in folder_name):

            # Found the week folder, now look for squeal folder inside
            cycles_audio_folder = week_folder / squeal_name

            if not cycles_audio_folder.exists():
                return None

            # Find most recent .cycles file
            cycles_files = list(cycles_audio_folder.glob("*.cycles"))
            if not cycles_files:
                return None

            # Read cycles from most recent file
            cycles_file = sorted(cycles_files)[-1]
            try:
                cycles = io.read_cycles(str(cycles_file))
                return len(cycles)
            except:
                return None

    return None


def consolidate_features_for_audio(audio_folder, squeal_name, subject, treatment_type, week, cycles_folder):
    """
    Consolidate all feature files for a single audio into one row

    Parameters:
    -----------
    audio_folder : Path
        Path to folder containing pickle files for this audio
    squeal_name : str
        Audio file name
    subject : str
        Subject name
    treatment_type : str
        Treatment type (B/S/U)
    week : int
        Week number
    cycles_folder : Path
        Base cycles folder

    Returns:
    --------
    pd.DataFrame : Single row with all features, or None if no features found
    """

    # Metadata
    metadata = {
        'Squeal': squeal_name,
        'Subject': subject,
        'Treatment': treatment_type,
        'Week': week,
        'Num_Cycles': get_num_cycles(cycles_folder, squeal_name, treatment_type, subject, week)
    }

    # Load all feature files
    all_features = []
    feature_sources = []  # Track which extractor each column came from

    # Define order of feature files
    feature_files_order = [
        'a_gat.pickle',
        'b_dfg.pickle',
        'c_temporal.pickle',
        'd_spectral.pickle',
        'e1_mfcc.pickle',
        'e2_gfcc.pickle',
        'e3_plpcc.pickle',
        'e4_rastacc.pickle',
        'e5_pncc.pickle',
        'f_lpc.pickle',
        'g1_pert.pickle',
        'g2_ms.pickle',
        'g3_comp_fix.pickle',
        'h_imf.pickle',
        'i_ratio.pickle',
        'z_ucla_parameters.pickle',
        'z2_schlegel21_parameters.pickle',
    ]

    for filename in feature_files_order:
        pickle_file = audio_folder / filename

        if pickle_file.exists():
            df_feat = load_features_from_pickle(pickle_file)

            if df_feat is not None:
                all_features.append(df_feat)
                # Track source for each column
                extractor_name = filename.replace('.pickle', '')
                for col in df_feat.columns:
                    feature_sources.append(extractor_name)

    if not all_features:
        return None, None

    # Combine all features
    features_df = pd.concat(all_features, axis=1)

    # Add metadata at the beginning
    for key, value in reversed(list(metadata.items())):
        features_df.insert(0, key, value)

    # Add metadata to feature_sources (prepend with 'metadata')
    feature_sources = ['metadata'] * len(metadata) + feature_sources

    return features_df, feature_sources


def apply_color_coding(excel_file, feature_sources):
    """
    Apply color coding to Excel headers based on feature source

    Parameters:
    -----------
    excel_file : str or Path
        Path to Excel file
    feature_sources : list
        List of source extractors for each column
    """

    # Load workbook
    wb = load_workbook(excel_file)
    ws = wb.active

    # Apply colors to header row (row 1)
    for col_idx, source in enumerate(feature_sources, start=1):
        cell = ws.cell(row=1, column=col_idx)

        if source == 'metadata':
            # Metadata columns - use a distinct color (light gray)
            fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        else:
            # Feature columns - use extractor-specific color
            color = EXTRACTOR_COLORS.get(source, 'FFFFFF')  # Default to white if not found
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        cell.fill = fill

    # Save workbook
    wb.save(excel_file)
    print(f"  Applied color coding to headers")


def get_parameter_description(param_name):
    """
    Get description for a parameter name using pattern matching

    Parameters:
    -----------
    param_name : str
        Parameter name

    Returns:
    --------
    str : Description of the parameter
    """
    param_lower = param_name.lower()

    # Check exact match first
    if param_name in PARAMETER_DESCRIPTIONS:
        return PARAMETER_DESCRIPTIONS[param_name]
    if param_lower in PARAMETER_DESCRIPTIONS:
        return PARAMETER_DESCRIPTIONS[param_lower]

    # Pattern matching for common parameter types
    desc_parts = []

    # Cepstral coefficients
    if 'mfcc' in param_lower:
        desc_parts.append('Mel-frequency cepstral coefficient')
    elif 'gfcc' in param_lower:
        desc_parts.append('Gammatone-frequency cepstral coefficient')
    elif 'plp' in param_lower or 'plpcc' in param_lower:
        desc_parts.append('Perceptual linear prediction coefficient')
    elif 'rasta' in param_lower:
        desc_parts.append('RASTA-PLP coefficient')
    elif 'pncc' in param_lower:
        desc_parts.append('Power-normalized cepstral coefficient')
    elif 'lpcc' in param_lower:
        desc_parts.append('Linear prediction cepstral coefficient')
    elif 'lpc' in param_lower:
        desc_parts.append('Linear prediction coefficient')
    elif 'lsf' in param_lower:
        desc_parts.append('Line spectral frequency')

    # Statistics
    if 'delta2' in param_lower or 'deltadelta' in param_lower:
        desc_parts.append('(acceleration/2nd derivative)')
    elif 'delta' in param_lower:
        desc_parts.append('(velocity/1st derivative)')

    if '_mean' in param_lower or 'mean' in param_lower:
        desc_parts.append('- mean across frames')
    elif '_std' in param_lower or 'std' in param_lower:
        desc_parts.append('- standard deviation')
    elif '_min' in param_lower:
        desc_parts.append('- minimum value')
    elif '_max' in param_lower:
        desc_parts.append('- maximum value')

    # Voice quality
    if 'jitter' in param_lower:
        return 'Jitter - cycle-to-cycle frequency variation'
    if 'shimmer' in param_lower:
        return 'Shimmer - cycle-to-cycle amplitude variation'
    if 'hnr' in param_lower:
        return 'Harmonics-to-noise ratio (dB)'
    if 'nhr' in param_lower:
        return 'Noise-to-harmonics ratio'
    if 'snr' in param_lower:
        return 'Signal-to-noise ratio (dB)'
    if 'cpp' in param_lower:
        return 'Cepstral peak prominence (dB)'
    if 'f0' in param_lower or 'pitch' in param_lower:
        return 'Fundamental frequency / pitch measure'

    # Spectral
    if 'centroid' in param_lower:
        return 'Spectral centroid - center of mass of spectrum'
    if 'bandwidth' in param_lower:
        return 'Spectral bandwidth - spread around centroid'
    if 'rolloff' in param_lower:
        return 'Spectral rolloff frequency'
    if 'flux' in param_lower:
        return 'Spectral flux - frame-to-frame change'
    if 'flatness' in param_lower:
        return 'Spectral flatness (0=tonal, 1=noisy)'
    if 'contrast' in param_lower:
        return 'Spectral contrast - peak-valley difference'

    # Formants
    if 'formant' in param_lower or param_lower.startswith('f1') or param_lower.startswith('f2') or param_lower.startswith('f3'):
        return 'Formant frequency (vocal tract resonance)'

    # Energy/amplitude
    if 'energy' in param_lower:
        return 'Signal energy measure'
    if 'amplitude' in param_lower:
        return 'Signal amplitude measure'
    if 'zcr' in param_lower:
        return 'Zero-crossing rate'

    # AVCA/complexity
    if 'rpde' in param_lower:
        return 'Recurrence period density entropy'
    if 'dfa' in param_lower:
        return 'Detrended fluctuation analysis'
    if 'ppe' in param_lower:
        return 'Pitch period entropy'
    if 'entropy' in param_lower:
        return 'Entropy measure (signal complexity)'

    # IMF
    if 'imf' in param_lower:
        return 'Intrinsic mode function from EMD'

    # Ratios
    if 'alpha' in param_lower and 'ratio' in param_lower:
        return 'Alpha ratio - low/high frequency energy ratio'
    if 'hammar' in param_lower:
        return 'Hammarberg index - spectral slope measure'
    if 'slope' in param_lower:
        return 'Spectral slope measure'
    if 'tilt' in param_lower:
        return 'Spectral tilt - energy distribution'

    # UCLA parameters
    if param_lower == 'ucla_f0':
        return 'Fundamental frequency (Hz) via autocorrelation (UCLA method)'
    if param_lower == 'ucla_jitter':
        return 'Jitter - period-to-period F0 variation ratio (UCLA method)'
    if param_lower == 'ucla_shimmer':
        return 'Shimmer - amplitude variation between periods ratio (UCLA method)'
    if param_lower == 'ucla_q50':
        return 'Q50 - spectral centroid at 50th percentile of cumulative energy (Hz)'
    if param_lower == 'ucla_flux':
        return 'Spectral flux - mean frame-to-frame spectral change (UCLA method)'

    # Schlegel21 parameters
    if param_lower == 'schlegel21_q50_2':
        return 'Q50 at call start — first of 9 partial windows (Tallet 2013, window 2/11)'
    if param_lower == 'schlegel21_q50_10':
        return 'Q50 at call end — last of 9 partial windows (Tallet 2013, window 10/11)'
    if param_lower == 'schlegel21_q50_min':
        return 'Minimum Q50 across 9 partial windows (Tallet 2013)'
    if param_lower == 'schlegel21_pf':
        return 'Peak frequency — frequency of maximum energy in whole-signal spectrum (Tallet 2013)'
    if param_lower == 'schlegel21_q25':
        return 'Q25 — mean 25th-percentile spectral energy frequency across frames (Garcia 2015)'
    if param_lower == 'schlegel21_dur':
        return 'Signal duration in seconds'
    if param_lower == 'schlegel21_hnr':
        return 'Harmonics-to-noise ratio via Praat cross-correlation (Linhart 2015), minimum_pitch=50 Hz'

    # If we found some description parts, join them
    if desc_parts:
        return ' '.join(desc_parts)

    # Default
    return 'Audio feature parameter'


def create_parameter_reference_sheet(excel_file, extracted_features_folder):
    """
    Create a parameter reference sheet with all parameters organized by extractor

    Parameters:
    -----------
    excel_file : str or Path
        Path to Excel file
    extracted_features_folder : Path
        Folder containing extracted features to read parameter names from
    """
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    print("\nCreating parameter reference sheet...")

    # Load workbook
    wb = load_workbook(excel_file)

    # Create new sheet for parameter reference
    if 'Parameter_Reference' in wb.sheetnames:
        del wb['Parameter_Reference']
    ws_ref = wb.create_sheet('Parameter_Reference')

    # Create new sheet for summary
    if 'Summary' in wb.sheetnames:
        del wb['Summary']
    ws_summary = wb.create_sheet('Summary')

    # Find a sample audio folder to read parameter names from
    sample_audio_folder = None
    for treatment_folder in extracted_features_folder.iterdir():
        if treatment_folder.is_dir():
            for subject_folder in treatment_folder.iterdir():
                if subject_folder.is_dir():
                    for week_folder in subject_folder.iterdir():
                        if week_folder.is_dir():
                            for audio_folder in week_folder.iterdir():
                                if audio_folder.is_dir():
                                    sample_audio_folder = audio_folder
                                    break
                            if sample_audio_folder:
                                break
                    if sample_audio_folder:
                        break
            if sample_audio_folder:
                break

    if not sample_audio_folder:
        print("  Warning: Could not find sample audio folder for parameter names")
        wb.save(excel_file)
        return

    # Collect parameters from each extractor
    extractor_params = {}
    feature_files_order = [
        'a_gat.pickle', 'b_dfg.pickle', 'c_temporal.pickle', 'd_spectral.pickle',
        'e1_mfcc.pickle', 'e2_gfcc.pickle', 'e3_plpcc.pickle', 'e4_rastacc.pickle',
        'e5_pncc.pickle', 'f_lpc.pickle', 'g1_pert.pickle', 'g2_ms.pickle',
        'g3_comp_fix.pickle', 'h_imf.pickle', 'i_ratio.pickle', 'z_ucla_parameters.pickle',
    ]

    for filename in feature_files_order:
        pickle_file = sample_audio_folder / filename
        extractor_name = filename.replace('.pickle', '')

        if pickle_file.exists():
            try:
                with open(pickle_file, 'rb') as f:
                    df = pickle.load(f)
                extractor_params[extractor_name] = list(df.columns)
            except Exception as e:
                print(f"  Warning: Could not read {filename}: {e}")
                extractor_params[extractor_name] = []
        else:
            extractor_params[extractor_name] = []

    # Styles
    header_font = Font(bold=True, size=11)
    subheader_font = Font(bold=True, size=10)
    normal_font = Font(size=9)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # =========================================================================
    # PARAMETER REFERENCE SHEET
    # =========================================================================

    # Write headers for each extractor (3 columns per extractor: Parameter, Description, empty)
    col = 1
    extractor_columns = {}  # Track which column each extractor starts at

    for extractor_name in feature_files_order:
        extractor_key = extractor_name.replace('.pickle', '')
        params = extractor_params.get(extractor_key, [])

        if not params:
            continue

        extractor_columns[extractor_key] = col

        # Get extractor info
        ext_info = EXTRACTOR_FULL_DESCRIPTIONS.get(extractor_key, {})
        ext_full_name = ext_info.get('name', extractor_key)
        ext_description = ext_info.get('description', '')
        ext_reference = ext_info.get('reference', '')

        # Extractor header (merged across 2 columns)
        cell = ws_ref.cell(row=1, column=col, value=ext_full_name)
        cell.font = header_font
        cell.alignment = center_align
        color = EXTRACTOR_COLORS.get(extractor_key, 'FFFFFF')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.border = thin_border

        # Also color the second column header
        cell2 = ws_ref.cell(row=1, column=col+1, value='')
        cell2.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell2.border = thin_border
        ws_ref.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)

        # Sub-headers
        cell_param = ws_ref.cell(row=2, column=col, value='Parameter')
        cell_param.font = subheader_font
        cell_param.alignment = center_align
        cell_param.border = thin_border

        cell_desc = ws_ref.cell(row=2, column=col+1, value='Description')
        cell_desc.font = subheader_font
        cell_desc.alignment = center_align
        cell_desc.border = thin_border

        # Write parameters
        for row_idx, param in enumerate(params, start=3):
            # Parameter name
            cell_p = ws_ref.cell(row=row_idx, column=col, value=param)
            cell_p.font = normal_font
            cell_p.alignment = left_align
            cell_p.border = thin_border

            # Description
            desc = get_parameter_description(param)
            cell_d = ws_ref.cell(row=row_idx, column=col+1, value=desc)
            cell_d.font = normal_font
            cell_d.alignment = left_align
            cell_d.border = thin_border

        # Set column widths using get_column_letter (avoids MergedCell issue)
        ws_ref.column_dimensions[get_column_letter(col)].width = 25
        ws_ref.column_dimensions[get_column_letter(col + 1)].width = 45

        col += 3  # Move to next extractor (2 columns + 1 gap)

    # =========================================================================
    # SUMMARY SHEET
    # =========================================================================

    # Title
    ws_summary.cell(row=1, column=1, value='FEATURE EXTRACTION SUMMARY').font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:E1')

    # Summary table headers
    headers = ['Extractor', 'Full Name', 'Parameters', 'Description', 'Reference']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.border = thin_border

    # Summary data
    total_params = 0
    row = 4
    for extractor_name in feature_files_order:
        extractor_key = extractor_name.replace('.pickle', '')
        params = extractor_params.get(extractor_key, [])
        num_params = len(params)

        if num_params == 0:
            continue

        total_params += num_params

        ext_info = EXTRACTOR_FULL_DESCRIPTIONS.get(extractor_key, {})
        ext_full_name = ext_info.get('name', extractor_key)
        ext_description = ext_info.get('description', '')
        ext_reference = ext_info.get('reference', '')

        # Extractor name with color
        cell = ws_summary.cell(row=row, column=1, value=extractor_key)
        cell.font = normal_font
        cell.alignment = center_align
        color = EXTRACTOR_COLORS.get(extractor_key, 'FFFFFF')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.border = thin_border

        # Full name
        cell = ws_summary.cell(row=row, column=2, value=ext_full_name)
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = thin_border

        # Number of parameters
        cell = ws_summary.cell(row=row, column=3, value=num_params)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border

        # Description
        cell = ws_summary.cell(row=row, column=4, value=ext_description)
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = thin_border

        # Reference
        cell = ws_summary.cell(row=row, column=5, value=ext_reference)
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = thin_border

        row += 1

    # Total row
    cell = ws_summary.cell(row=row, column=1, value='TOTAL')
    cell.font = Font(bold=True)
    cell.alignment = center_align
    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    cell.border = thin_border

    ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell2 = ws_summary.cell(row=row, column=2)
    cell2.border = thin_border

    cell = ws_summary.cell(row=row, column=3, value=total_params)
    cell.font = Font(bold=True)
    cell.alignment = center_align
    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    cell.border = thin_border

    # Empty cells for description and reference in total row
    for col_idx in [4, 5]:
        cell = ws_summary.cell(row=row, column=col_idx, value='')
        cell.border = thin_border

    # Set column widths for summary sheet
    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 40
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 80
    ws_summary.column_dimensions['E'].width = 35

    # Rename first sheet
    wb.worksheets[0].title = 'Features_Data'

    # Save workbook
    wb.save(excel_file)
    print(f"  Created 'Parameter_Reference' sheet with parameter details")
    print(f"  Created 'Summary' sheet with {total_params} total parameters across {len([k for k, v in extractor_params.items() if v])} extractors")


def consolidate_all_features(extracted_features_folder, cycles_folder, output_excel_path):
    """
    Consolidate all extracted features into a single Excel file

    Parameters:
    -----------
    extracted_features_folder : Path
        Base folder containing extracted features
    cycles_folder : Path
        Base cycles folder
    output_excel_path : Path
        Output Excel file path

    Returns:
    --------
    pd.DataFrame : Consolidated features DataFrame
    """

    print("="*70)
    print("CONSOLIDATING EXTRACTED FEATURES")
    print("="*70)
    print(f"\nInput folder:  {extracted_features_folder}")
    print(f"Output Excel:  {output_excel_path}")
    print()

    all_rows = []
    feature_sources_list = None  # Will store column sources from first audio

    # Traverse the folder structure: Treatment/Subject/Week/Audio
    treatment_folders = [f for f in extracted_features_folder.iterdir() if f.is_dir()]

    # Count total audio folders for progress bar
    total_audio_folders = 0
    for treatment_folder in treatment_folders:
        for subject_folder in treatment_folder.iterdir():
            if subject_folder.is_dir():
                for week_folder in subject_folder.iterdir():
                    if week_folder.is_dir():
                        total_audio_folders += len([f for f in week_folder.iterdir() if f.is_dir()])

    total_processed = 0

    # Create progress bar
    pbar = tqdm(total=total_audio_folders,
                desc="Consolidating features",
                unit="file",
                ncols=100)

    for treatment_folder in sorted(treatment_folders):
        treatment_name = treatment_folder.name

        # Map treatment name back to code
        treatment_code_map = {
            'Bilateral': 'B',
            'Scar': 'S',
            'Unilateral': 'U'
        }
        treatment_code = treatment_code_map.get(treatment_name, '?')

        subject_folders = [f for f in treatment_folder.iterdir() if f.is_dir()]

        for subject_folder in sorted(subject_folders):
            subject_name = subject_folder.name

            week_folders = [f for f in subject_folder.iterdir() if f.is_dir()]

            for week_folder in sorted(week_folders):
                # Extract week number from folder name (e.g., "week_0" -> 0)
                try:
                    week_num = int(week_folder.name.split('_')[1])
                except:
                    continue

                audio_folders = [f for f in week_folder.iterdir() if f.is_dir()]

                for audio_folder in sorted(audio_folders):
                    squeal_name = audio_folder.name

                    # Update progress bar description
                    pbar.set_description(f"{treatment_code} | {subject_name[:15]:15s} | Week {week_num:2d}")

                    # Consolidate features for this audio
                    row_df, feature_sources = consolidate_features_for_audio(
                        audio_folder, squeal_name, subject_name,
                        treatment_code, week_num, cycles_folder
                    )

                    if row_df is not None:
                        all_rows.append(row_df)

                        # Store feature sources from first audio
                        if feature_sources_list is None:
                            feature_sources_list = feature_sources

                        total_processed += 1

                    pbar.update(1)

    pbar.close()
    print(f"\nTotal audio files processed: {total_processed}")

    if not all_rows:
        print("ERROR: No features found to consolidate!")
        return None

    # Combine all rows into single DataFrame
    print("\nCombining all features...")
    consolidated_df = pd.concat(all_rows, ignore_index=True)

    # Drop columns with more than 50% missing values (these add noise, not information)
    METADATA_COLS = {'Squeal', 'Subject', 'Treatment', 'Week', 'Num_Cycles'}
    NAN_THRESHOLD = 0.50
    nan_fraction = consolidated_df.isna().mean()
    cols_to_drop = [c for c in nan_fraction[nan_fraction > NAN_THRESHOLD].index
                    if c not in METADATA_COLS]
    if cols_to_drop:
        print(f"\nDropping {len(cols_to_drop)} columns with >{NAN_THRESHOLD*100:.0f}% missing values:")
        for c in cols_to_drop:
            print(f"  - {c} ({nan_fraction[c]*100:.1f}% NaN)")
        consolidated_df = consolidated_df.drop(columns=cols_to_drop)

    # Drop columns with known systematic quality issues identified via audit:
    #   f0_max    — quantization artifacts: 94.7% of rows are integer-period boundary values
    #   f0_min    — 97.3% of values fall below the valid 50 Hz F0 floor for pig vocalizations
    #   snr1_max  — hard cap at 36 dB inside svfel noise.snr_v1_gat(); 27.2% of rows hit it
    #   snr1_min  — 5.8% extreme outliers (near-zero values, >10× IQR below Q1)
    KNOWN_BAD_COLS = ['f0_max', 'f0_min', 'snr1_max', 'snr1_min']
    to_drop_known = [c for c in KNOWN_BAD_COLS if c in consolidated_df.columns]
    if to_drop_known:
        print(f"\nDropping {len(to_drop_known)} columns with known systematic quality issues:")
        for c in to_drop_known:
            print(f"  - {c}")
        consolidated_df = consolidated_df.drop(columns=to_drop_known)

    # Rebuild feature_sources_list to match remaining columns
    if feature_sources_list is not None:
        # Build a column→source map from the original sources list
        # We need the original column order before dropping — reconstruct from first row
        original_cols = list(all_rows[0].columns) if all_rows else []
        col_to_source = dict(zip(original_cols, feature_sources_list[:len(original_cols)]))
        feature_sources_list = [col_to_source.get(c, 'metadata') for c in consolidated_df.columns]

    # Sort by Treatment, Subject, Week
    print("\nSorting by Treatment, Subject, Week...")
    consolidated_df = consolidated_df.sort_values(
        by=['Treatment', 'Subject', 'Week'],
        ignore_index=True
    )

    # Save to Excel
    print(f"\nSaving to Excel: {output_excel_path}")
    consolidated_df.to_excel(output_excel_path, index=False, engine='openpyxl')

    # Apply color coding
    print("Applying color coding to headers...")
    apply_color_coding(output_excel_path, feature_sources_list)

    # Create parameter reference and summary sheets
    create_parameter_reference_sheet(output_excel_path, extracted_features_folder)

    print("\n" + "="*70)
    print("CONSOLIDATION COMPLETE!")
    print("="*70)
    print(f"\nTotal rows: {len(consolidated_df)}")
    print(f"Total columns: {len(consolidated_df.columns)}")
    print(f"  - Metadata columns: 5 (Squeal, Subject, Treatment, Week, Num_Cycles)")
    print(f"  - Feature columns: {len(consolidated_df.columns) - 5}")
    print(f"\nOutput file: {output_excel_path}")

    # Show color legend
    print("\n" + "="*70)
    print("COLOR LEGEND")
    print("="*70)
    print("Metadata columns: Gray")
    print("\nFeature columns by extractor:")

    extractor_descriptions = {
        'a_gat': 'GAT features (37): F0, jitter, shimmer, HNR, SNR, etc.',
        'b_dfg': 'DFG features (14): CPP, CPPS, harmonic measures',
        'c_temporal': 'Temporal features (6): Amplitude, energy, ZCR',
        'd_spectral': 'Spectral features (30): Centroid, bandwidth, contrast, etc.',
        'e1_mfcc': 'MFCC features (72): MFCCs with deltas',
        'e2_gfcc': 'GFCC features (72): Gammatone coefficients with deltas',
        'e3_plpcc': 'PLPCC features (72): PLP coefficients with deltas',
        'e4_rastacc': 'RASTA-PLP features (72): RASTA-PLP with deltas',
        'e5_pncc': 'PNCC features (72): Power-normalized cepstral coefficients',
        'f_lpc': 'LPC features (64): LPC, LSF, LPCC',
        'g1_pert': 'AVCA Perturbation features',
        'g2_ms': 'AVCA Modulation spectrum features',
        'g3_comp_fix': 'AVCA Complexity features',
        'h_imf': 'IMF features',
        'i_ratio': 'Ratio features (4): Spectral ratios',
        'z_ucla_parameters': 'UCLA parameters (5): F0, Jitter, Shimmer, Q50, Flux',
        'z2_schlegel21_parameters': 'Schlegel21 parameters (7): Q50_2, Q50_10, Q50_min, PF, Q25, Dur, HNR'
    }

    for extractor, color in EXTRACTOR_COLORS.items():
        desc = extractor_descriptions.get(extractor, extractor)
        print(f"  {extractor:15s} (#{color}): {desc}")

    return consolidated_df


def main():
    """Main function"""

    # Configuration
    # BASE_PATH is the repository's Feature_Extraction/ folder (parent of this
    # scripts/ directory), so the layout matches main_pig_directory_scan_final.py:
    #   Feature_Extraction/Extracted_Features_Structured/
    #   Feature_Extraction/Detected_Cycles/
    #   Feature_Extraction/Consolidated_Features_final.xlsx
    script_dir = Path(__file__).parent
    BASE_PATH = script_dir.parent

    EXTRACTED_FEATURES_FOLDER = BASE_PATH / "Extracted_Features_Structured"
    CYCLES_FOLDER = BASE_PATH / "Detected_Cycles"  # Now mirrors Audio_Snips structure
    OUTPUT_EXCEL = BASE_PATH / "Consolidated_Features_final.xlsx"

    # Check if extracted features folder exists
    if not EXTRACTED_FEATURES_FOLDER.exists():
        print(f"ERROR: Extracted features folder not found: {EXTRACTED_FEATURES_FOLDER}")
        print("\nPlease run main_pig_directory_scan_final.py first to extract features.")
        return

    # Add svfel to path for cycles reading
    sys.path.append(str(script_dir))
    sys.path.append(str(script_dir / "svfel"))

    # Consolidate features
    df = consolidate_all_features(EXTRACTED_FEATURES_FOLDER, CYCLES_FOLDER, OUTPUT_EXCEL)

    if df is not None:
        print("\n✓ Feature consolidation successful!")
        print(f"\nYou can now open: {OUTPUT_EXCEL}")
        print("Headers are color-coded by feature extractor source.")


if __name__ == '__main__':
    main()
