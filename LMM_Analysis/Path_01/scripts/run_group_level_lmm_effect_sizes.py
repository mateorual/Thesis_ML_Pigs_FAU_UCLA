"""
LMM Group-Level Effect Size Analysis — 59 Acoustic Features × 3 Treatment Groups
==================================================================================
Fits: feature ~ Stage + (1 | Subject)   [categorical, REML, Kenward-Roger df]
Reference level: Pre-Surgery
Input:  Selected_Features_Groups_S_U_B_v1.xlsx
Outputs: effect_size_LMM_results_group_level.xlsx (9 sheets)

Methodological blueprint: unified_parameter_analysis_B_Klomhaus_v2.py
"""

import os
import sys

# Windows: configure R_HOME and DLL search path before rpy2/pymer4 are loaded.
# Prefer an existing R_HOME; otherwise use the newest R installation in Program Files.
if sys.platform == 'win32':
    _r_home = os.environ.get('R_HOME')
    if not _r_home:
        _r_root = os.path.join(os.environ.get('ProgramFiles', r'C:\Program Files'), 'R')
        if os.path.isdir(_r_root):
            _candidates = [
                os.path.join(_r_root, name)
                for name in os.listdir(_r_root)
                if os.path.isdir(os.path.join(_r_root, name))
            ]
            if _candidates:
                _r_home = sorted(_candidates)[-1]
                os.environ['R_HOME'] = _r_home

    if _r_home and os.path.isdir(_r_home):
        _r_bin = os.path.join(_r_home, 'bin', 'x64')
        os.environ['PATH'] = _r_bin + os.pathsep + os.environ.get('PATH', '')
        if hasattr(os, 'add_dll_directory') and os.path.isdir(_r_bin):
            os.add_dll_directory(_r_bin)

import logging
import warnings
import numpy as np
import pandas as pd
from scipy.stats import t as t_dist
from statsmodels.stats.multitest import multipletests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

# ============================================================================
# PYMER4 IMPORT (Kenward-Roger correction — required, no fallback)
# ============================================================================
from pymer4.models import Lmer

# ============================================================================
# CONFIGURATION
# ============================================================================

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(_SCRIPT_DIR)
INPUT_FILE  = os.path.join(PROJECT_DIR, "data", "Selected_Features_Groups_S_U_B_v1.xlsx")
INPUT_SHEET = "Features_Data"

OUTPUT_DIR  = os.path.join(PROJECT_DIR, "outputs", "generated")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "effect_size_LMM_results_group_level.xlsx")
os.makedirs(OUTPUT_DIR, exist_ok=True)

METADATA_COLS = ['Squeal', 'Subject', 'Treatment', 'Week', 'Num_Cycles', 'Stage']

# Treatment group membership (Subject column values)
TREATMENT_SUBJECTS = {
    'B': ['Beck and Kurt', 'Frank and Dean', 'Madonna and Beyonce', 'Taylor and Gaga'],
    'U': ['Michael and Prince', 'Paul and John', 'Tina and Aretha'],
    'S': ['Barry and Stevie', 'Cher and Adele', 'Elvis', 'Snoop and Dre'],
}

STAGE_ORDER    = ['Pre-Surgery', 'Early-Post', 'Late-Post']
STAGE_REFERENCE = 'Pre-Surgery'

# Stage value mapping: dataset labels → internal names ('Mid Post' is dropped)
STAGE_MAP = {
    'Pre':        'Pre-Surgery',
    'Early Post': 'Early-Post',
    'Late Post':  'Late-Post',
    'Mid Post':   None,
}


CONTRAST_NAMES = {
    'PreVsEarly':   ('Pre-Surgery', 'Early-Post'),
    'PreVsLate':    ('Pre-Surgery', 'Late-Post'),
    'EarlyVsLate':  ('Early-Post',  'Late-Post'),
}

# Sheet naming: {group}_{contrast_key}
SHEET_ORDER = [
    'LMM_B_PreVsEarly', 'LMM_B_PreVsLate', 'LMM_B_EarlyVsLate',
    'LMM_U_PreVsEarly', 'LMM_U_PreVsLate', 'LMM_U_EarlyVsLate',
    'LMM_S_PreVsEarly', 'LMM_S_PreVsLate', 'LMM_S_EarlyVsLate',
    'Summary',
]

# Note: features are Z-scored before fitting, so Beta_Raw == Cohen_d by construction.
# Both columns are retained for traceability.
# Beta_Raw: raw LMM coefficient on Z-scored feature (equivalent to Cohen's d when feature is standardized).
DETAIL_COLS = [
    'Feature', 'Treatment', 'Contrast', 'n_obs', 'n_subjects',
    'Beta_Raw', 'SE', 'df_KR', 't_statistic', 'p_value', 'p_value_adj',
    'Cohen_d', 'Interpretation', 'CI_Lower', 'CI_Upper',
    'R2_Marginal', 'R2_Conditional', 'Converged', 'Low_Power_Warning',
]

# Excel fill colours
FILLS = {
    'Negligible': PatternFill('solid', fgColor='FFFFFFFF'),
    'Small':      PatternFill('solid', fgColor='FFFFFF99'),
    'Medium':     PatternFill('solid', fgColor='FFFFCC66'),
    'Large':      PatternFill('solid', fgColor='FFFF9999'),
    'LowPower':   PatternFill('solid', fgColor='FFD6EAF8'),
}

# ============================================================================
# LOGGING
# ============================================================================

def setup_logging(output_dir):
    log_file = os.path.join(output_dir, 'lmm_analysis_warnings.log')
    logger = logging.getLogger('lmm_analysis')
    logger.setLevel(logging.DEBUG)
    logger.handlers = []

    fh = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    fh.setLevel(logging.WARNING)
    fh.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(fh)

    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter('%(message)s'))
    logger.addHandler(ch)

    return logger


logger = None   # initialised in main()

# ============================================================================
# HELPERS
# ============================================================================

def get_t_critical(df_degrees, alpha=0.05):
    if np.isnan(df_degrees) or df_degrees <= 0:
        df_degrees = 1
    return t_dist.ppf(1 - alpha / 2, df_degrees)


def cohen_d_interpretation(d):
    abs_d = abs(d)
    if abs_d < 0.2:
        return 'Negligible'
    elif abs_d < 0.5:
        return 'Small'
    elif abs_d < 0.8:
        return 'Medium'
    else:
        return 'Large'

# ============================================================================
# LMM FITTING — single feature, single treatment group
# ============================================================================

def fit_lmm_for_feature(df_group, feature, group_label, feat_idx=0):
    """
    Fit: feature ~ Stage + (1 | Subject)
    Returns a dict with per-contrast statistics (3 contrasts).
    """
    result = {
        'converged': False,
        'r2_marginal': np.nan,
        'r2_conditional': np.nan,
        'contrasts': {},   # key: contrast_key → stats dict
        'outcome_sd': np.nan,
        'n_obs': 0,
        'n_subjects': 0,
    }

    # CORRECTION 4: n_obs / n_subjects set AFTER dropna to avoid overcounting
    df_work = df_group[['Stage', 'Subject', feature]].dropna().copy()
    result['n_obs'] = len(df_work)
    result['n_subjects'] = df_work['Subject'].nunique()

    if len(df_work) < 5:
        logger.warning(f"[{group_label}] {feature}: too few observations ({len(df_work)}), skipping.")
        return result

    # CORRECTION 1: Z-score standardization — Beta_Raw from model equals Cohen's d directly
    outcome_mean = df_work[feature].mean()
    outcome_sd   = df_work[feature].std()
    result['outcome_sd'] = outcome_sd

    if outcome_sd == 0 or np.isnan(outcome_sd):
        logger.warning(f"[{group_label}] {feature}: zero/NaN SD, skipping.")
        return result

    df_work[feature] = (df_work[feature] - outcome_mean) / outcome_sd

    # CORRECTION 4: log unbalanced pig-stage combinations
    stage_subject_counts = df_work.groupby('Stage')['Subject'].nunique()
    for stage, count in stage_subject_counts.items():
        if count < result['n_subjects']:
            logger.warning(
                f"[{group_label}] {feature}: Stage '{stage}' has only {count}/"
                f"{result['n_subjects']} subjects (unbalanced data)."
            )

    # CORRECTION 3: collision-safe column name using feature index
    safe_feat = f"feat_{feat_idx}"
    df_work[safe_feat] = df_work[feature]

    # Set Stage as unordered Categorical with explicit level order, ref = first level
    df_work['Stage'] = pd.Categorical(
        df_work['Stage'],
        categories=STAGE_ORDER,
        ordered=False
    )
    assert df_work['Stage'].cat.categories[0] == STAGE_REFERENCE, \
        f"Reference level must be '{STAGE_REFERENCE}' (first category). Check STAGE_ORDER."

    formula = f"{safe_feat} ~ Stage + (1 | Subject)"

    try:
        model = Lmer(formula, data=df_work)
        model.fit(REML=True)

        result['converged'] = True

        # R² — Nakagawa marginal/conditional (manual, pymer4 0.8.x has no .r2 attr)
        # var_fixed  = variance of design_matrix @ beta  (fixed-effect linear predictor)
        # var_random = sum of non-residual variance components from ranef_var
        # var_resid  = residual variance from ranef_var
        try:
            X      = np.array(model.design_matrix)
            beta   = model.coefs['Estimate'].values
            X_beta = X @ beta
            var_fixed  = float(np.var(X_beta))
            rv         = model.ranef_var
            var_random = float(rv.loc[rv.index != 'Residual', 'Var'].sum())
            var_resid  = float(rv.loc['Residual', 'Var'])
            var_total  = var_fixed + var_random + var_resid
            if var_total > 0:
                result['r2_marginal']    = float(np.clip(var_fixed / var_total, 0, 1))
                result['r2_conditional'] = float(np.clip((var_fixed + var_random) / var_total, 0, 1))
            else:
                result['r2_marginal']    = np.nan
                result['r2_conditional'] = np.nan
        except Exception as e:
            logger.warning(f"[{group_label}] {feature}: R² extraction failed: {e}")

        coef_df = model.coefs

        # ---- extract vcov via rpy2 for proper SE on EarlyVsLate ----
        # Uses globalenv assignment + r('as.matrix(vcov(...))') because rpy2 runs in
        # ABI mode on Windows and cannot convert S4 lmerMod objects via localconverter.
        vcov = None
        vcov_index = None
        try:
            if hasattr(model, 'model_obj') and model.model_obj is not None:
                from rpy2 import robjects as _robj
                _robj.globalenv['.tmp_vcov_model'] = model.model_obj
                vcov_mat   = _robj.r('as.matrix(vcov(.tmp_vcov_model))')
                vcov_names = _robj.r('rownames(vcov(.tmp_vcov_model))')
                vcov       = np.array(vcov_mat)
                vcov_index = list(vcov_names)
        except Exception as e:
            logger.warning(f"[{group_label}] {feature}: vcov extraction failed: {e}")

        # ---- helper: find a coefficient row by trying several naming patterns ----
        def find_coef_row(stage_name):
            patterns = [
                f"Stage[T.{stage_name}]",
                f"StageT.{stage_name}",
                f"Stage{stage_name}",
                stage_name.replace('-', '_').replace(' ', '_'),
                stage_name,
            ]
            for p in patterns:
                if p in coef_df.index:
                    return coef_df.loc[p], p
            return None, None

        # ---- helper: find vcov index by stage name ----
        def find_vcov_idx(stage_name):
            if vcov_index is None:
                return None
            patterns = [
                f"Stage[T.{stage_name}]",
                f"StageT.{stage_name}",
                f"Stage{stage_name}",
                stage_name.replace('-', '_').replace(' ', '_'),
                stage_name,
            ]
            for p in patterns:
                if p in vcov_index:
                    return vcov_index.index(p)
            return None

        # ---- extract direct contrasts (Early-Post vs Pre, Late-Post vs Pre) ----
        def extract_direct_contrast(stage_name):
            row, matched_name = find_coef_row(stage_name)
            if row is None:
                logger.warning(f"[{group_label}] {feature}: coefficient not found for Stage={stage_name}. "
                                f"Available: {coef_df.index.tolist()}")
                return None
            raw    = float(row['Estimate'])
            se     = float(row['SE'])
            df_kr  = float(row.get('DF', 1))
            t_stat = float(row.get('T-stat', raw / se if se > 0 else np.nan))
            pval   = float(row.get('P-val', np.nan))
            t_crit = get_t_critical(df_kr)
            cohen_d = raw  # feature is Z-scored; Beta_Raw == Cohen's d directly
            return {
                'Beta_Raw': raw, 'SE': se, 'df_KR': df_kr,
                't_statistic': t_stat, 'p_value': pval,
                'CI_Lower': raw - t_crit * se,
                'CI_Upper': raw + t_crit * se,
                'Cohen_d': cohen_d,
                '_matched_name': matched_name,
            }

        early_stats = extract_direct_contrast('Early-Post')
        late_stats  = extract_direct_contrast('Late-Post')

        if early_stats is not None:
            result['contrasts']['PreVsEarly'] = early_stats
        if late_stats is not None:
            result['contrasts']['PreVsLate'] = late_stats

        # ---- Late-Post vs Early-Post: difference of contrasts ----
        if early_stats is not None and late_stats is not None:
            estimate = late_stats['Beta_Raw'] - early_stats['Beta_Raw']

            # Attempt proper SE via vcov
            se_from_cov = False
            if vcov is not None and vcov_index is not None:
                try:
                    idx_early = find_vcov_idx('Early-Post')
                    idx_late  = find_vcov_idx('Late-Post')
                    if idx_early is not None and idx_late is not None:
                        var_e  = vcov[idx_early, idx_early]
                        var_l  = vcov[idx_late,  idx_late]
                        cov_el = vcov[idx_early, idx_late]
                        var_diff = var_e + var_l - 2 * cov_el
                        if var_diff > 0:
                            se_diff = np.sqrt(var_diff)
                            se_from_cov = True
                except Exception as e:
                    logger.warning(f"[{group_label}] {feature}: vcov SE for EarlyVsLate failed: {e}")

            if not se_from_cov:
                # Independence-assumption fallback (conservative)
                se_diff = np.sqrt(early_stats['SE']**2 + late_stats['SE']**2)
                logger.info(f"[{group_label}] {feature}: EarlyVsLate using independence-assumption SE.")

            df_diff  = min(early_stats['df_KR'], late_stats['df_KR'])
            t_diff   = estimate / se_diff if se_diff > 0 else np.nan
            pval_diff = (2 * (1 - t_dist.cdf(abs(t_diff), df_diff))
                         if not np.isnan(t_diff) else np.nan)
            t_crit_diff = get_t_critical(df_diff)
            cohen_d_diff = estimate  # Z-scored feature; estimate == Cohen's d directly

            result['contrasts']['EarlyVsLate'] = {
                'Beta_Raw': estimate, 'SE': se_diff, 'df_KR': df_diff,
                't_statistic': t_diff, 'p_value': pval_diff,
                'CI_Lower': estimate - t_crit_diff * se_diff,
                'CI_Upper': estimate + t_crit_diff * se_diff,
                'Cohen_d': cohen_d_diff,
            }

    except Exception as e:
        result['converged'] = False
        logger.warning(f"[{group_label}] {feature}: LMM failed: {e}")

    return result


# ============================================================================
# RUN ALL FEATURES FOR ONE TREATMENT GROUP
# ============================================================================

def run_group(df_stage, group_label, subjects, features, low_power=False):
    """
    Fits LMMs for all features in one treatment group.
    Returns dict: contrast_key → list of row dicts (one per feature).
    """
    df_grp = df_stage[df_stage['Subject'].isin(subjects)].copy()

    if low_power:
        print(f"\n{'='*70}")
        print(f"WARNING - Group {group_label}: n={df_grp['Subject'].nunique()} subjects. "
              "p-values are unreliable with few subjects; "
              "effect sizes should be interpreted with caution.")
        print(f"{'='*70}\n")

    all_rows = {ck: [] for ck in CONTRAST_NAMES}
    n_features = len(features)

    for i, feat in enumerate(features):
        if (i + 1) % 25 == 0 or (i + 1) == n_features:
            print(f"[LMM - Treatment {group_label}] Feature {i+1}/{n_features}: {feat}")

        res = fit_lmm_for_feature(df_grp, feat, group_label, feat_idx=i)

        for ck, (ref_stage, comp_stage) in CONTRAST_NAMES.items():
            c = res['contrasts'].get(ck, {})
            beta   = c.get('Beta_Raw',    np.nan)
            cohen  = c.get('Cohen_d',     np.nan)
            interp = cohen_d_interpretation(cohen) if not np.isnan(cohen) else np.nan

            row = {
                'Feature':          feat,
                'Treatment':        group_label,
                'Contrast':         f"{comp_stage} vs {ref_stage}",
                'n_obs':            res['n_obs'],
                'n_subjects':       res['n_subjects'],
                'Beta_Raw':         beta,
                'SE':               c.get('SE',          np.nan),
                'df_KR':            c.get('df_KR',        np.nan),
                't_statistic':      c.get('t_statistic',  np.nan),
                'p_value':          c.get('p_value',      np.nan),
                'p_value_adj':      np.nan,   # filled after FDR correction
                'Cohen_d':          cohen,
                'Interpretation':   interp,
                'CI_Lower':         c.get('CI_Lower',     np.nan),
                'CI_Upper':         c.get('CI_Upper',     np.nan),
                'R2_Marginal':      res['r2_marginal'],
                'R2_Conditional':   res['r2_conditional'],
                'Converged':        res['converged'],
                'Low_Power_Warning': low_power,
            }
            all_rows[ck].append(row)

    # ---- FDR correction across all 3 contrasts × all features ----
    # CORRECTION 2: build flat_rows once — same order for collection and assignment
    flat_rows   = [row for ck in CONTRAST_NAMES for row in all_rows[ck]]
    all_pvals   = [row['p_value'] for row in flat_rows]
    valid_mask  = [not np.isnan(p) for p in all_pvals]
    pvals_valid = [p for p, m in zip(all_pvals, valid_mask) if m]

    if pvals_valid:
        _, padj_valid, _, _ = multipletests(pvals_valid, method='fdr_bh')
        padj_iter = iter(padj_valid)
        for row, is_valid in zip(flat_rows, valid_mask):
            row['p_value_adj'] = float(next(padj_iter)) if is_valid else np.nan
    else:
        logger.warning(f"[{group_label}] No valid p-values for FDR correction.")

    return all_rows


# ============================================================================
# EXCEL OUTPUT
# ============================================================================

def _auto_width(ws, max_width=40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col_cells
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _interp_fill(value):
    return FILLS.get(value, FILLS['Negligible'])


def write_detail_sheet(ws, rows_df, low_power):
    # Header
    ws.append(DETAIL_COLS)
    ws.freeze_panes = "A2"

    interp_col_idx = DETAIL_COLS.index('Interpretation') + 1   # 1-based

    for _, row in rows_df.iterrows():
        data = [row.get(c, np.nan) for c in DETAIL_COLS]
        ws.append(data)

        xl_row = ws.max_row
        interp_val = row.get('Interpretation', '')

        if low_power:
            # Highlight entire row in light blue
            row_fill = FILLS['LowPower']
            for cell in ws[xl_row]:
                cell.fill = row_fill
        else:
            # Colour only the Interpretation cell
            interp_cell = ws.cell(row=xl_row, column=interp_col_idx)
            if isinstance(interp_val, str) and interp_val in FILLS:
                interp_cell.fill = _interp_fill(interp_val)

    _auto_width(ws)


def write_summary_sheet(ws, summary_rows):
    cols = [
        'Treatment', 'Contrast', 'Total_Features', 'Converged',
        'Negligible', 'Small', 'Medium', 'Large',
        'Significant_p_adj_0.05',
    ]
    ws.append(cols)
    ws.freeze_panes = "A2"
    for r in summary_rows:
        ws.append([r.get(c, 0) for c in cols])
    _auto_width(ws)


def build_summary_rows(all_group_rows):
    """Build one summary row per (group, contrast)."""
    summary = []
    for group_label in ['B', 'U', 'S']:
        for ck in CONTRAST_NAMES:
            rows = all_group_rows.get((group_label, ck), [])
            df = pd.DataFrame(rows)
            if df.empty:
                continue
            ref_stage, comp_stage = CONTRAST_NAMES[ck]
            contrast_str = f"{comp_stage} vs {ref_stage}"
            summary.append({
                'Treatment':                group_label,
                'Contrast':                 contrast_str,
                'Total_Features':           len(df),
                'Converged':                int(df['Converged'].sum()),
                'Negligible':               int((df['Interpretation'] == 'Negligible').sum()),
                'Small':                    int((df['Interpretation'] == 'Small').sum()),
                'Medium':                   int((df['Interpretation'] == 'Medium').sum()),
                'Large':                    int((df['Interpretation'] == 'Large').sum()),
                'Significant_p_adj_0.05':   int((df['p_value_adj'] < 0.05).sum()),
            })
    return summary


# ============================================================================
# CHECKPOINT HELPERS
# ============================================================================

# Reverse map: contrast string → contrast key  (built from CONTRAST_NAMES)
_CONTRAST_STR_TO_KEY = {
    f"{comp} vs {ref}": ck
    for ck, (ref, comp) in CONTRAST_NAMES.items()
}


def _checkpoint_path(group_label):
    return os.path.join(OUTPUT_DIR, f"checkpoint_{group_label}.csv")


def save_checkpoint(group_label, all_rows):
    """
    Flatten all contrast rows for this group into one CSV and save it.
    Called immediately after run_group() completes so progress is not lost.
    """
    frames = []
    for ck in CONTRAST_NAMES:
        rows = all_rows.get(ck, [])
        if rows:
            frames.append(pd.DataFrame(rows, columns=DETAIL_COLS))
    if not frames:
        return
    df_cp = pd.concat(frames, ignore_index=True)
    path = _checkpoint_path(group_label)
    df_cp.to_csv(path, index=False)
    print(f"  [Checkpoint] Saved {len(df_cp)} rows -> {os.path.basename(path)}")


def load_checkpoint(group_label):
    """
    Load a previously saved checkpoint CSV for this group.
    Returns dict contrast_key → list[row_dict], or None if no checkpoint exists.
    """
    path = _checkpoint_path(group_label)
    if not os.path.exists(path):
        return None
    df_cp = pd.read_csv(path)
    # Restore bool columns that CSV reads as object/float
    for col in ('Converged', 'Low_Power_Warning'):
        if col in df_cp.columns:
            df_cp[col] = df_cp[col].map(
                lambda v: True if str(v).strip().lower() in ('true', '1') else False
            )
    all_rows = {ck: [] for ck in CONTRAST_NAMES}
    for _, row in df_cp.iterrows():
        ck = _CONTRAST_STR_TO_KEY.get(str(row.get('Contrast', '')))
        if ck is not None:
            all_rows[ck].append(row.to_dict())
    n_loaded = sum(len(v) for v in all_rows.values())
    print(f"  [Checkpoint] Loaded {n_loaded} rows from {os.path.basename(path)} — skipping refit.")
    return all_rows


# ============================================================================
# MAIN
# ============================================================================

def main():
    # Force UTF-8 on Windows console to avoid cp1252 encoding errors in print()
    if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

    global logger
    logger = setup_logging(OUTPUT_DIR)
    logger.info(f"LMM Group-Level Analysis — started {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # ---- Load data ----
    print(f"\nLoading: {INPUT_FILE}")
    df = pd.read_excel(INPUT_FILE, sheet_name=INPUT_SHEET)
    print(f"  Loaded {len(df)} rows, {len(df.columns)} columns")

    # ---- Remap Stage values to internal names, drop Mid Post ----
    df['Stage'] = df['Stage'].map(STAGE_MAP)
    n_before = len(df)
    df = df[df['Stage'].notna()].copy()
    n_dropped = n_before - len(df)
    if n_dropped:
        print(f"  Dropped {n_dropped} rows (Mid Post / unmapped stages).")

    # ---- Identify feature columns ----
    features = [c for c in df.columns if c not in METADATA_COLS]
    print(f"  Features: {len(features)}")
    print(f"  Stages: {df['Stage'].value_counts().to_dict()}")

    # ---- Keep df_stage for subsetting ----
    df_stage = df.copy()

    # ---- Fit models per group (with checkpoint resume) ----
    all_group_rows = {}   # (group_label, contrast_key) → list of row dicts

    for group_label, subjects in TREATMENT_SUBJECTS.items():
        low_power = (group_label == 'U')
        print(f"\n{'='*70}")
        print(f"Processing Treatment Group {group_label} "
              f"({len(subjects)} subjects): {subjects}")
        print(f"{'='*70}")

        # Resume from checkpoint if one exists for this group
        cached = load_checkpoint(group_label)
        if cached is not None:
            rows_by_contrast = cached
        else:
            rows_by_contrast = run_group(
                df_stage, group_label, subjects, features, low_power=low_power
            )
            save_checkpoint(group_label, rows_by_contrast)

        for ck, rows in rows_by_contrast.items():
            all_group_rows[(group_label, ck)] = rows

    # ---- Write Excel ----
    print(f"\nWriting results to: {OUTPUT_FILE}")
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    for group_label in ['B', 'U', 'S']:
        low_power = (group_label == 'U')
        for ck in CONTRAST_NAMES:
            sheet_name = f"LMM_{group_label}_{ck}"
            rows = all_group_rows.get((group_label, ck), [])
            ws = wb.create_sheet(title=sheet_name)
            df_sheet = pd.DataFrame(rows, columns=DETAIL_COLS)
            write_detail_sheet(ws, df_sheet, low_power)
            print(f"  Sheet '{sheet_name}': {len(df_sheet)} rows")

    # Summary sheet
    summary_rows = build_summary_rows(all_group_rows)
    ws_sum = wb.create_sheet(title='Summary')
    write_summary_sheet(ws_sum, summary_rows)

    wb.save(OUTPUT_FILE)
    print(f"\n[DONE] Saved: {OUTPUT_FILE}")
    logger.info(f"Analysis complete. Output: {OUTPUT_FILE}")


if __name__ == '__main__':
    main()
