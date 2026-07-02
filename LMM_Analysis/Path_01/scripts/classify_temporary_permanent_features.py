"""
permanent_temporary_features_group_level.py

Classifies features per treatment group (B, U, S) into Temporary, Permanent,
or Other based on LMM effect size interpretations across key stage contrasts
from effect_size_LMM_results_group_level.xlsx.
Produces temporary_permanent_group_level.xlsx.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR   = os.path.dirname(SCRIPT_DIR)
GENERATED_DIR = os.path.join(PROJECT_DIR, "outputs", "generated")
REFERENCE_DIR = os.path.join(PROJECT_DIR, "outputs", "reference_results")
os.makedirs(GENERATED_DIR, exist_ok=True)

GENERATED_INPUT = os.path.join(GENERATED_DIR, "effect_size_LMM_results_group_level.xlsx")
REFERENCE_INPUT = os.path.join(REFERENCE_DIR, "effect_size_LMM_results_group_level.xlsx")
INPUT_FILE = GENERATED_INPUT if os.path.exists(GENERATED_INPUT) else REFERENCE_INPUT
OUTPUT_FILE = os.path.join(GENERATED_DIR, "temporary_permanent_group_level.xlsx")

# ── Constants ──────────────────────────────────────────────────────────────────
HIGH = {"Large"}
LOW  = {"Negligible", "Small"}

GROUPS = ("B", "U", "S")

# Color fills
FILL_TEMPORARY = PatternFill("solid", fgColor="FFD6EAF8")   # light blue
FILL_PERMANENT = PatternFill("solid", fgColor="FFFF9999")   # light red
FILL_U_ROW     = PatternFill("solid", fgColor="FFFFFF99")   # light yellow
FILL_HEADER    = PatternFill("solid", fgColor="FFD9D9D9")   # light grey


# ── 1. Load data ───────────────────────────────────────────────────────────────
def load_group(group: str) -> dict[str, pd.DataFrame]:
    """Load PreVsEarly, PreVsLate, EarlyVsLate sheets for one group."""
    sheets = {}
    for contrast_key, suffix in [
        ("PreVsEarly", "PreVsEarly"),
        ("PreVsLate",  "PreVsLate"),
        ("EarlyVsLate","EarlyVsLate"),
    ]:
        sheet_name = f"LMM_{group}_{suffix}"
        sheets[contrast_key] = pd.read_excel(INPUT_FILE, sheet_name=sheet_name)
    return sheets


# ── 2. Classify features for one group ────────────────────────────────────────
def classify_group(group: str) -> pd.DataFrame:
    sheets = load_group(group)

    pe  = sheets["PreVsEarly"]   # Pre vs Early
    pl  = sheets["PreVsLate"]    # Pre vs Late
    el  = sheets["EarlyVsLate"]  # Early vs Late (supplementary)

    # Rename columns we need from each sheet
    pe_cols = pe[["Feature", "Cohen_d", "Interpretation", "p_value_adj",
                   "R2_Marginal", "R2_Conditional", "Converged", "Low_Power_Warning"]].rename(columns={
        "Cohen_d":         "Cohen_d_PreEarly",
        "Interpretation":  "Interp_PreEarly",
        "p_value_adj":     "p_value_adj_PreEarly",
        "R2_Marginal":     "R2_Marginal_PreEarly",
        "R2_Conditional":  "R2_Conditional_PreEarly",
        "Converged":       "Converged_PreEarly",
        "Low_Power_Warning": "Low_Power_Warning",
    })

    pl_cols = pl[["Feature", "Cohen_d", "Interpretation", "p_value_adj",
                   "Converged"]].rename(columns={
        "Cohen_d":        "Cohen_d_PreLate",
        "Interpretation": "Interp_PreLate",
        "p_value_adj":    "p_value_adj_PreLate",
        "Converged":      "Converged_PreLate",
    })

    el_cols = el[["Feature", "Cohen_d", "Interpretation"]].rename(columns={
        "Cohen_d":        "Cohen_d_EarlyLate",
        "Interpretation": "Interp_EarlyLate",
    })

    # Merge all three on Feature
    merged = pe_cols.merge(pl_cols, on="Feature", how="outer")
    merged = merged.merge(el_cols, on="Feature", how="left")

    # Fill missing convergence as False (absent = not converged)
    merged["Converged_PreEarly"] = merged["Converged_PreEarly"].fillna(False)
    merged["Converged_PreLate"]  = merged["Converged_PreLate"].fillna(False)

    merged["Treatment"] = group

    def categorize(row):
        # Must have converged on both key contrasts
        if not row["Converged_PreEarly"] or not row["Converged_PreLate"]:
            return "Other"
        ie = row["Interp_PreEarly"]
        il = row["Interp_PreLate"]
        if ie in HIGH and il in LOW:
            return "Temporary"
        elif ie in HIGH and il in HIGH:
            return "Permanent"
        else:
            return "Other"

    merged["Category"] = merged.apply(categorize, axis=1)

    # Sort: Category then descending |Cohen_d_PreEarly| then Feature
    merged["_abs_cohend"] = merged["Cohen_d_PreEarly"].abs()
    merged = merged.sort_values(
        ["Category", "_abs_cohend", "Feature"],
        ascending=[True, False, True],
    ).drop(columns=["_abs_cohend"]).reset_index(drop=True)

    return merged


# ── 3. Detail sheet columns ────────────────────────────────────────────────────
DETAIL_COLS = [
    "Feature", "Treatment", "Category",
    "Cohen_d_PreEarly", "Interp_PreEarly",
    "Cohen_d_PreLate",  "Interp_PreLate",
    "Cohen_d_EarlyLate","Interp_EarlyLate",
    "p_value_adj_PreEarly", "p_value_adj_PreLate",
    "R2_Marginal_PreEarly", "R2_Conditional_PreEarly",
    "Converged_PreEarly", "Converged_PreLate",
    "Low_Power_Warning",
]


# ── 4. Summary sheet ───────────────────────────────────────────────────────────
def compute_summary(classified: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for group in GROUPS:
        df = classified[group]
        converged_both = (df["Converged_PreEarly"] & df["Converged_PreLate"]).sum()
        rows.append({
            "Treatment":                  group,
            "Total_Features_Tested":      len(df),
            "Converged_Both_Contrasts":   converged_both,
            "Temporary":                  (df["Category"] == "Temporary").sum(),
            "Permanent":                  (df["Category"] == "Permanent").sum(),
            "Other":                      (df["Category"] == "Other").sum(),
            "Low_Power_Warning":          df["Low_Power_Warning"].sum() if "Low_Power_Warning" in df.columns else 0,
        })
    return pd.DataFrame(rows)


# ── 5. Cross-group consensus sheet ────────────────────────────────────────────
def compute_cross_group(classified: dict[str, pd.DataFrame]) -> pd.DataFrame:
    # Build a lookup: feature -> category per group
    cat_map = {}
    cohend_map = {}
    for group in GROUPS:
        df = classified[group]
        for _, row in df.iterrows():
            feat = row["Feature"]
            if feat not in cat_map:
                cat_map[feat]   = {}
                cohend_map[feat] = {}
            cat_map[feat][group]    = row["Category"]
            cohend_map[feat][group] = row["Cohen_d_PreEarly"]

    rows = []
    for feat, cats in cat_map.items():
        for target_cat in ("Temporary", "Permanent"):
            agreeing = [g for g in GROUPS if cats.get(g) == target_cat]
            if len(agreeing) >= 2:
                # Low-power warning for U
                u_low_power = False
                if "U" in agreeing:
                    u_df = classified["U"]
                    u_row = u_df[u_df["Feature"] == feat]
                    if len(u_row) and u_row.iloc[0].get("Low_Power_Warning", False):
                        u_low_power = True

                rows.append({
                    "Feature":            feat,
                    "Category":           target_cat,
                    "Groups":             ", ".join(agreeing),
                    "n_groups_agreeing":  len(agreeing),
                    "Category_B":         cats.get("B", "—"),
                    "Category_U":         cats.get("U", "—"),
                    "Category_S":         cats.get("S", "—"),
                    "Cohen_d_PreEarly_B": cohend_map[feat].get("B", np.nan),
                    "Cohen_d_PreEarly_U": cohend_map[feat].get("U", np.nan),
                    "Cohen_d_PreEarly_S": cohend_map[feat].get("S", np.nan),
                    "U_Low_Power":        u_low_power,
                })

    if not rows:
        return pd.DataFrame(columns=[
            "Feature", "Category", "Groups", "n_groups_agreeing",
            "Category_B", "Category_U", "Category_S",
            "Cohen_d_PreEarly_B", "Cohen_d_PreEarly_U", "Cohen_d_PreEarly_S",
            "U_Low_Power",
        ])

    result = pd.DataFrame(rows)
    result = result.sort_values(
        ["Category", "n_groups_agreeing", "Feature"],
        ascending=[True, False, True],
    ).drop(columns=["n_groups_agreeing"]).reset_index(drop=True)
    return result


# ── 6. Excel formatting ────────────────────────────────────────────────────────
def format_sheet(ws, df: pd.DataFrame,
                 category_col: str = None,
                 is_group_u: bool = False,
                 u_warning_note: bool = False):
    col_names = list(df.columns)

    # Header row
    for cell in ws[1]:
        cell.font      = Font(bold=True)
        cell.fill      = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Optional U-warning note in the first unused column header
    if u_warning_note:
        note_col_idx = len(col_names) + 1
        note_cell = ws.cell(1, note_col_idx)
        note_cell.value     = "Group U: n=3 pig pairs; interpret p-values with caution"
        note_cell.font      = Font(bold=True, color="FF8B0000")
        note_cell.fill      = FILL_HEADER
        note_cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(note_col_idx)].width = 45

    ws.freeze_panes = "A2"

    # Auto-width (capped at 45)
    for col_idx, col in enumerate(col_names, start=1):
        lengths = [len(str(col))] + [len(str(v)) for v in df[col].fillna("")]
        max_len = max(lengths)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    cat_idx = (col_names.index(category_col) + 1) if category_col and category_col in col_names else None

    # Data rows
    for row_idx in range(2, ws.max_row + 1):
        # Group U background
        if is_group_u:
            for col_idx in range(1, len(col_names) + 1):
                ws.cell(row_idx, col_idx).fill = FILL_U_ROW

        # Category cell fill (overwrites U fill for that cell)
        if cat_idx:
            val = ws.cell(row_idx, cat_idx).value
            if val == "Temporary":
                ws.cell(row_idx, cat_idx).fill = FILL_TEMPORARY
            elif val == "Permanent":
                ws.cell(row_idx, cat_idx).fill = FILL_PERMANENT


# ── 7. Main ────────────────────────────────────────────────────────────────────
def main():
    print("Loading and classifying data...")
    classified = {g: classify_group(g) for g in GROUPS}

    summary    = compute_summary(classified)
    cross_grp  = compute_cross_group(classified)

    # Detail: Temporary + Permanent only
    detail = {
        g: classified[g][classified[g]["Category"].isin(["Temporary", "Permanent"])][DETAIL_COLS].copy()
        for g in GROUPS
    }

    print(f"Writing output to {OUTPUT_FILE} ...")
    sheet_defs = []  # (sheet_name, df, category_col, is_U, u_warning_note)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for group in GROUPS:
            sname = f"GroupLevel_{group}"
            detail[group].to_excel(writer, sheet_name=sname, index=False)
            sheet_defs.append((sname, detail[group], "Category",
                                group == "U", group == "U"))

        summary.to_excel(writer, sheet_name="Summary", index=False)
        sheet_defs.append(("Summary", summary, None, False, False))

        cross_grp.to_excel(writer, sheet_name="CrossGroup_Consensus", index=False)
        sheet_defs.append(("CrossGroup_Consensus", cross_grp, "Category", False, False))

    # Post-processing: formatting
    wb = load_workbook(OUTPUT_FILE)
    for (sname, df, cat_col, is_u, u_note) in sheet_defs:
        ws = wb[sname]
        format_sheet(ws, df, category_col=cat_col,
                     is_group_u=is_u, u_warning_note=u_note)
    wb.save(OUTPUT_FILE)

    print("Done.\n")

    # ── Quick stats ──────────────────────────────────────────────────────────
    print("=== Quick Stats ===")
    print(summary.to_string(index=False))
    print(f"\nCross-group consensus rows: {len(cross_grp)}")
    if len(cross_grp):
        print(cross_grp[["Feature", "Category", "Groups"]].to_string(index=False))


if __name__ == "__main__":
    main()
